import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.models import (
    Quote, RawMaterial, MouldingMachineDetail, Assembly,
    AssemblyRawMaterial, ManufacturingPrintingCost, Packaging, Transport
)
import tempfile
import os


class ExcelTemplateGenerator:
    """Generate Excel templates for uploading data"""

    @staticmethod
    def _add_vertical_headers(ws, headers, color, start_row=1):
        """Add vertical headers in column A with color coding"""
        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for row_num, header in enumerate(headers, start_row):
            cell = ws.cell(row=row_num, column=1)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.column_dimensions['A'].width = 35

    @staticmethod
    def create_raw_materials_template():
        """Create template for raw materials upload (vertical format)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Raw Materials"

        headers = [
            'Material Name*',
            'Grade',
            'RM Code*',
            'Unit (kg/gm/ton)*',
            'RM Rate (per kg)*',
            'Frozen Rate (per kg)',
            'Part Weight*',
            'Runner Weight*',
            'Process Losses %',
            'Purging Losses %',
            'Other RM Cost',
            'Other RM Cost Description',
            'ICC %',
            'Rejection %',
            'Overhead %',
            'Maintenance %',
            'Profit %'
        ]

        ExcelTemplateGenerator._add_vertical_headers(ws, headers, "4472C4")

        # Add sample data for 3 materials
        materials = [
            ['PP Copolymer', 'PP-H340R', 'RM-PP-001', 'kg', 125.50, None, 0.0234, 0.0045, 2.5, 1.2, 5.0, 'Masterbatch cost', 0.5, 2.0, 5.0, 3.0, 10.0],
            ['ABS High Impact', 'ABS-750', 'RM-ABS-002', 'kg', 185.75, 180.00, 0.0456, 0.0089, 3.0, 1.5, 7.5, 'Color additive', 0.75, 2.5, 4.5, 2.5, 12.0],
            ['Nylon 6', 'PA6-GF30', 'RM-PA6-003', 'kg', 215.00, None, 0.0678, 0.0123, 4.5, 2.0, 10.0, 'UV stabilizer', 1.0, 3.0, 6.0, 4.0, 15.0],
        ]

        for col_num, material_data in enumerate(materials, 2):
            for row_num, value in enumerate(material_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
            ws.column_dimensions[get_column_letter(col_num)].width = 18

        return wb

    @staticmethod
    def create_moulding_machines_template():
        """Create template for moulding machines upload (vertical format)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Moulding Machines"

        headers = [
            'Cavity*',
            'Machine Tonnage*',
            'Cycle Time (s)*',
            'Efficiency %*',
            'Shift Rate*',
            'Shift Rate for MTC*',
            'MTC Count*',
            'Rejection %',
            'Overhead %',
            'Maintenance %',
            'Profit %'
        ]

        ExcelTemplateGenerator._add_vertical_headers(ws, headers, "70AD47")

        # Add sample data for 3 machines
        machines = [
            [2, 120, 25.5, 85.0, 2500.00, 2200.00, 500, 1.0, 5.0, 3.0, 10.0],
            [4, 180, 28.3, 87.5, 3200.00, 2800.00, 1000, 1.2, 5.2, 3.2, 11.0],
            [8, 250, 30.8, 90.0, 4500.00, 4000.00, 2000, 0.8, 4.8, 2.8, 12.0],
        ]

        for col_num, machine_data in enumerate(machines, 2):
            for row_num, value in enumerate(machine_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        return wb

    @staticmethod
    def create_assemblies_template():
        """Create template for assemblies upload (vertical format)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Assemblies"

        headers = [
            'Assembly Name*',
            'Assembly Type',
            'Remarks',
            'Manual Cost',
            'Other Cost',
            'Other Cost Description',
            'Inspection & Handling Cost',
            'Profit %',
            'Rejection %'
        ]

        ExcelTemplateGenerator._add_vertical_headers(ws, headers, "FFC000")

        # Add sample data
        assemblies = [
            ['Manual Screw Assembly', 'Manual', 'M3 screws with washers', 8.50, 2.50, 'Screw materials', 5.00, 15.0, 2.0],
            ['Ultrasonic Welding', 'Automated', 'High frequency welding', 12.00, 3.50, 'Welding consumables', 8.00, 18.0, 1.5],
        ]

        for col_num, assembly_data in enumerate(assemblies, 2):
            for row_num, value in enumerate(assembly_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        return wb

    @staticmethod
    def create_packaging_template():
        """Create template for packaging upload (vertical format)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Packaging"

        headers = [
            'Packaging Category* (box/polybag)',
            'Parts per Packaging*',
            'Maintenance %',
            '--- BOX FIELDS ---',
            'Length (mm)',
            'Breadth (mm)',
            'Height (mm)',
            'Cost',
            'Lifecycle',
            '--- POLYBAG FIELDS ---',
            'Polybag Length (inches)',
            'Polybag Width (inches)',
            'Rate per kg',
            'Polybags per kg'
        ]

        ExcelTemplateGenerator._add_vertical_headers(ws, headers, "E26B0A")

        # Add sample data for 2 packaging options (1 box, 1 polybag)
        packagings = [
            ['box', 50, 5.0, '', 600, 400, 250, 350.00, 100, '', 0, 0, 0, 0],
            ['polybag', 100, 3.0, '', 0, 0, 0, 0, 0, '', 16, 20, 250.00, 1000],
        ]

        for col_num, pkg_data in enumerate(packagings, 2):
            for row_num, value in enumerate(pkg_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
            ws.column_dimensions[get_column_letter(col_num)].width = 18

        return wb

    @staticmethod
    def create_transport_template():
        """Create template for transport upload (vertical format)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Transport"

        headers = [
            'Length (ft)*',
            'Breadth (ft)*',
            'Height (ft)*',
            'Trip Cost*',
            'Parts per Box*'
        ]

        ExcelTemplateGenerator._add_vertical_headers(ws, headers, "9933FF")

        # Add sample data
        transports = [
            [12, 8, 6, 6500.00, 1000],
            [14, 9, 7, 8500.00, 1500],
        ]

        for col_num, trans_data in enumerate(transports, 2):
            for row_num, value in enumerate(trans_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        return wb

    @staticmethod
    def create_complete_quote_template():
        """Create template for complete quote with all sections"""
        wb = Workbook()
        wb.remove(wb.active)

        # Instructions Sheet
        ws_inst = wb.create_sheet("Instructions")
        ws_inst['A1'] = "COMPLETE QUOTE TEMPLATE"
        ws_inst['A1'].font = Font(bold=True, size=14)
        ws_inst['A3'] = "This template allows you to create a complete quote with all components:"
        ws_inst['A4'] = "- Quote definition"
        ws_inst['A5'] = "- Raw materials (RM Rate in per kg, Process/Purging Losses in %)"
        ws_inst['A6'] = "- Moulding machines"
        ws_inst['A7'] = "- Assemblies"
        ws_inst['A8'] = "- Packaging"
        ws_inst['A9'] = "- Transport (dimensions in feet)"
        ws_inst['A11'] = "IMPORTANT NOTES:"
        ws_inst['A11'].font = Font(bold=True)
        ws_inst['A12'] = "- RM Rate is per kg (converted to per gram for calculations)"
        ws_inst['A13'] = "- Process Losses and Purging Losses are percentages of (part weight + runner weight)"
        ws_inst['A14'] = "- Other RM Cost is an additional fixed cost added directly to the material cost"
        ws_inst['A15'] = "- Weights converted to grams: kg→×1000, ton→×1,000,000, gm→×1"
        ws_inst['A16'] = "- Transport dimensions are in feet (1 ft = 304.8 mm)"
        ws_inst['A17'] = "- Inspection & Handling Cost is a fixed cost (not percentage)"
        ws_inst.column_dimensions['A'].width = 80

        # Quote Definition
        ws_def = wb.create_sheet("Quote Definition")
        def_headers = [
            'Quote Name*',
            'Client Name*',
            'SAP Number',
            'Part Number*',
            'Part Name*',
            'Amendment Number',
            'Description',
            'Quantity*',
            'Handling Charge',
            'Profit %',
            'Notes'
        ]
        ExcelTemplateGenerator._add_vertical_headers(ws_def, def_headers, "2E75B6")

        quote_data = [
            'Sample Complete Quote',
            'ABC Manufacturing Ltd',
            'SAP-2024-001',
            'PART-12345',
            'Automotive Dashboard Component',
            'Rev-A',
            'Complete quote with all components for testing',
            10000,
            500.00,
            15.0,
            'Priority order - Q1 2024'
        ]
        for row_num, value in enumerate(quote_data, 1):
            ws_def.cell(row=row_num, column=2, value=value)
        ws_def.column_dimensions['B'].width = 40

        # Add other sheets
        wb_rm = ExcelTemplateGenerator.create_raw_materials_template()
        ws_rm = wb.create_sheet("Raw Materials")
        for row in wb_rm.active.iter_rows():
            for cell in row:
                ws_rm[cell.coordinate].value = cell.value
                if cell.has_style:
                    ws_rm[cell.coordinate].font = cell.font.copy()
                    ws_rm[cell.coordinate].fill = cell.fill.copy()
                    ws_rm[cell.coordinate].alignment = cell.alignment.copy()
        ws_rm.column_dimensions['A'].width = 35

        wb_mm = ExcelTemplateGenerator.create_moulding_machines_template()
        ws_mm = wb.create_sheet("Moulding Machines")
        for row in wb_mm.active.iter_rows():
            for cell in row:
                ws_mm[cell.coordinate].value = cell.value
                if cell.has_style:
                    ws_mm[cell.coordinate].font = cell.font.copy()
                    ws_mm[cell.coordinate].fill = cell.fill.copy()
                    ws_mm[cell.coordinate].alignment = cell.alignment.copy()
        ws_mm.column_dimensions['A'].width = 35

        wb_asm = ExcelTemplateGenerator.create_assemblies_template()
        ws_asm = wb.create_sheet("Assemblies")
        for row in wb_asm.active.iter_rows():
            for cell in row:
                ws_asm[cell.coordinate].value = cell.value
                if cell.has_style:
                    ws_asm[cell.coordinate].font = cell.font.copy()
                    ws_asm[cell.coordinate].fill = cell.fill.copy()
                    ws_asm[cell.coordinate].alignment = cell.alignment.copy()
        ws_asm.column_dimensions['A'].width = 35

        wb_pkg = ExcelTemplateGenerator.create_packaging_template()
        ws_pkg = wb.create_sheet("Packaging")
        for row in wb_pkg.active.iter_rows():
            for cell in row:
                ws_pkg[cell.coordinate].value = cell.value
                if cell.has_style:
                    ws_pkg[cell.coordinate].font = cell.font.copy()
                    ws_pkg[cell.coordinate].fill = cell.fill.copy()
                    ws_pkg[cell.coordinate].alignment = cell.alignment.copy()
        ws_pkg.column_dimensions['A'].width = 35

        wb_trans = ExcelTemplateGenerator.create_transport_template()
        ws_trans = wb.create_sheet("Transport")
        for row in wb_trans.active.iter_rows():
            for cell in row:
                ws_trans[cell.coordinate].value = cell.value
                if cell.has_style:
                    ws_trans[cell.coordinate].font = cell.font.copy()
                    ws_trans[cell.coordinate].fill = cell.fill.copy()
                    ws_trans[cell.coordinate].alignment = cell.alignment.copy()
        ws_trans.column_dimensions['A'].width = 35

        return wb

    @staticmethod
    def create_multiple_quotes_template():
        """Create template for uploading multiple complete quotes"""
        wb = Workbook()
        wb.remove(wb.active)

        # Instructions
        ws_inst = wb.create_sheet("Instructions")
        ws_inst['A1'] = "MULTIPLE COMPLETE QUOTES TEMPLATE"
        ws_inst['A1'].font = Font(bold=True, size=14)
        ws_inst['A3'] = "This file allows you to create multiple quotes with all components:"
        ws_inst['A4'] = "- Quote definition"
        ws_inst['A5'] = "- Raw materials (RM Rate in per kg)"
        ws_inst['A6'] = "- Moulding machines"
        ws_inst['A7'] = "- Assemblies"
        ws_inst['A8'] = "- Packaging"
        ws_inst['A9'] = "- Transport (dimensions in feet)"
        ws_inst['A11'] = "STRUCTURE:"
        ws_inst['A11'].font = Font(bold=True)
        ws_inst['A12'] = "- Each sheet uses vertical format (headers in column A)"
        ws_inst['A13'] = "- Each column (B, C, D...) represents one entry"
        ws_inst['A14'] = "- Link components to quotes using 'Quote Name' in row 1"
        ws_inst['A16'] = "IMPORTANT NOTES:"
        ws_inst['A16'].font = Font(bold=True)
        ws_inst['A17'] = "- RM Rate is entered per kg (will be converted to per gram for calculations)"
        ws_inst['A18'] = "- Process Losses and Purging Losses are percentages of (part weight + runner weight)"
        ws_inst['A19'] = "- Other RM Cost is an additional fixed cost added directly to the material cost"
        ws_inst['A20'] = "- Weights are converted to grams: kg→×1000, ton→×1,000,000, gm→×1"
        ws_inst['A21'] = "- Transport dimensions are in feet (1 ft = 304.8 mm)"
        ws_inst['A22'] = "- Inspection & Handling is a fixed cost (not percentage)"
        ws_inst['A24'] = "Upload this to a project to create multiple quotes with all components."
        ws_inst.column_dimensions['A'].width = 100

        # Quote Definition
        ws_def = wb.create_sheet("Quote Definition")
        def_headers = [
            'Quote Name*',
            'Client Name*',
            'SAP Number',
            'Part Number*',
            'Part Name*',
            'Amendment Number',
            'Description',
            'Quantity*',
            'Handling Charge',
            'Profit %',
            'Notes'
        ]
        ExcelTemplateGenerator._add_vertical_headers(ws_def, def_headers, "2E75B6")

        quotes = [
            ['Auto Dashboard 2024', 'AutoTech Industries', 'SAP-AUTO-2024-001', 'DASH-4567', 'Dashboard Center Console', 'Rev-B', 'Automotive dashboard component with premium finish', 10000, 500, 20, 'High volume production'],
            ['Electronics Housing Pro', 'TechCorp Ltd', 'SAP-TECH-2024-002', 'EH-PRO-890', 'Premium Electronics Housing', 'Rev-A', 'IP67 rated electronic component housing', 25000, 350, 18, 'Weather resistant coating required'],
            ['Medical Device Shell', 'MedEquip Co', 'SAP-MED-2024-003', 'MD-SHELL-123', 'Sterile Medical Device Shell', 'Rev-C', 'FDA compliant medical device component', 5000, 800, 25, 'Cleanroom manufacturing required'],
        ]

        for col_num, quote_data in enumerate(quotes, 2):
            for row_num, value in enumerate(quote_data, 1):
                ws_def.cell(row=row_num, column=col_num, value=value)
            ws_def.column_dimensions[get_column_letter(col_num)].width = 30

        # Raw Materials (horizontal format with Quote Name in row 1)
        ws_rm = wb.create_sheet("Raw Materials")
        rm_headers = [
            'Quote Name*',
            'Material Name*',
            'Grade',
            'RM Code*',
            'Unit (kg/gm/ton)*',
            'RM Rate (per kg)*',
            'Frozen Rate (per kg)',
            'Part Weight*',
            'Runner Weight*',
            'Process Losses %',
            'Purging Losses %',
            'Other RM Cost',
            'Other RM Cost Description',
            'ICC %',
            'Rejection %',
            'Overhead %',
            'Maintenance %',
            'Profit %'
        ]
        ExcelTemplateGenerator._add_vertical_headers(ws_rm, rm_headers, "4472C4")

        raw_materials = [
            ['Auto Dashboard 2024', 'PP Copolymer', 'PP-R272', 'RM-PP-272', 'kg', 135.5, None, 0.245, 0.038, 3.5, 1.8, 5.0, 'Masterbatch', 0.8, 2.2, 5.5, 3.2, 11],
            ['Auto Dashboard 2024', 'ABS High Impact', 'ABS-HI-850', 'RM-ABS-850', 'kg', 195, None, 0.156, 0.024, 2.8, 1.5, 7.5, 'UV stabilizer', 0.6, 2.0, 5.0, 3.0, 10],
            ['Electronics Housing Pro', 'PC+ABS Blend', 'PCABS-750', 'RM-PCABS-750', 'kg', 265, None, 0.178, 0.028, 3.2, 1.6, 8.0, 'Flame retardant', 0.7, 2.5, 6.0, 3.5, 12],
            ['Medical Device Shell', 'Medical Grade PP', 'PP-MED-H310', 'RM-PP-MED-310', 'kg', 385, None, 0.125, 0.019, 2, 1.2, 12.0, 'Antimicrobial', 0.5, 1.5, 4.0, 2.5, 8],
        ]

        for col_num, material in enumerate(raw_materials, 2):
            for row_num, value in enumerate(material, 1):
                ws_rm.cell(row=row_num, column=col_num, value=value)
            ws_rm.column_dimensions[get_column_letter(col_num)].width = 25

        # Moulding Machines (horizontal format)
        ws_mm = wb.create_sheet("Moulding Machines")
        mm_headers = [
            'Quote Name*',
            'Cavity*',
            'Machine Tonnage*',
            'Cycle Time (s)*',
            'Efficiency %*',
            'Shift Rate*',
            'Shift Rate for MTC*',
            'MTC Count*',
            'Rejection %',
            'Overhead %',
            'Maintenance %',
            'Profit %'
        ]
        ExcelTemplateGenerator._add_vertical_headers(ws_mm, mm_headers, "70AD47")

        machines = [
            ['Auto Dashboard 2024', 8, 320, 58.5, 88, 9500, 9000, 4, 2.8, 6.5, 4.2, 16],
            ['Auto Dashboard 2024', 4, 220, 52.3, 90, 7200, 6800, 3, 2.5, 6, 3.8, 14],
            ['Electronics Housing Pro', 4, 180, 45.7, 91, 6500, 6100, 3, 2.2, 5.5, 3.5, 13],
            ['Medical Device Shell', 2, 150, 42.8, 92.5, 5800, 5400, 2, 2, 5, 3.2, 12],
        ]

        for col_num, machine in enumerate(machines, 2):
            for row_num, value in enumerate(machine, 1):
                ws_mm.cell(row=row_num, column=col_num, value=value)
            ws_mm.column_dimensions[get_column_letter(col_num)].width = 25

        # Assemblies (horizontal format)
        ws_asm = wb.create_sheet("Assemblies")
        asm_headers = [
            'Quote Name*',
            'Assembly Name*',
            'Assembly Type',
            'Remarks',
            'Manual Cost',
            'Other Cost',
            'Other Cost Description',
            'Inspection & Handling Cost',
            'Profit %',
            'Rejection %'
        ]
        ExcelTemplateGenerator._add_vertical_headers(ws_asm, asm_headers, "FFC000")

        assemblies = [
            ['Auto Dashboard 2024', 'Insert Molding + Ultrasonic Welding', 'Automated', 'High precision required', 28.5, 6.5, 'Welding consumables', 10, 12, 3],
            ['Electronics Housing Pro', 'Snap Fit Assembly', 'Manual', 'IP67 sealing required', 15.75, 4.25, 'Sealing materials', 6.5, 10, 2.5],
            ['Medical Device Shell', 'Cleanroom Assembly + Sterilization', 'Manual', 'FDA Class II requirements', 45, 12, 'Sterilization', 15, 15, 1.5],
        ]

        for col_num, assembly in enumerate(assemblies, 2):
            for row_num, value in enumerate(assembly, 1):
                ws_asm.cell(row=row_num, column=col_num, value=value)
            ws_asm.column_dimensions[get_column_letter(col_num)].width = 30

        # Packaging (horizontal format)
        ws_pkg = wb.create_sheet("Packaging")
        pkg_headers = [
            'Quote Name*',
            'Packaging Category* (box/polybag)',
            'Parts per Packaging*',
            'Maintenance %',
            '--- BOX FIELDS ---',
            'Length (mm)',
            'Breadth (mm)',
            'Height (mm)',
            'Cost',
            'Lifecycle',
            '--- POLYBAG FIELDS ---',
            'Polybag Length (inches)',
            'Polybag Width (inches)',
            'Rate per kg',
            'Polybags per kg'
        ]
        ExcelTemplateGenerator._add_vertical_headers(ws_pkg, pkg_headers, "E26B0A")

        packagings = [
            ['Auto Dashboard 2024', 'box', 80, 4.5, '', 700, 450, 300, 420.00, 200, '', 0, 0, 0, 0],
            ['Electronics Housing Pro', 'box', 100, 3.8, '', 600, 400, 250, 350.00, 250, '', 0, 0, 0, 0],
            ['Medical Device Shell', 'polybag', 25, 2.5, '', 0, 0, 0, 0, 0, '', 14, 18, 280.00, 950],
        ]

        for col_num, pkg in enumerate(packagings, 2):
            for row_num, value in enumerate(pkg, 1):
                ws_pkg.cell(row=row_num, column=col_num, value=value)
            ws_pkg.column_dimensions[get_column_letter(col_num)].width = 25

        # Transport (horizontal format)
        ws_trans = wb.create_sheet("Transport")
        trans_headers = [
            'Quote Name*',
            'Length (ft)*',
            'Breadth (ft)*',
            'Height (ft)*',
            'Trip Cost*',
            'Parts per Box*'
        ]
        ExcelTemplateGenerator._add_vertical_headers(ws_trans, trans_headers, "9933FF")

        transports = [
            ['Auto Dashboard 2024', 14, 9, 7, 9500, 240],
            ['Electronics Housing Pro', 12, 8, 6, 7200, 300],
            ['Medical Device Shell', 10, 7, 5, 5500, 100],
        ]

        for col_num, trans in enumerate(transports, 2):
            for row_num, value in enumerate(trans, 1):
                ws_trans.cell(row=row_num, column=col_num, value=value)
            ws_trans.column_dimensions[get_column_letter(col_num)].width = 25

        return wb


class ExcelParser:
    """Parse Excel files and import data"""

    @staticmethod
    def parse_raw_materials(file_path, quote):
        """Parse raw materials from Excel (vertical format)"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Try different possible sheet names
        sheet_name = None
        for possible_name in ['Raw Materials', 'Raw_Materials', 'RawMaterials']:
            if possible_name in wb.sheetnames:
                sheet_name = possible_name
                break

        if not sheet_name:
            return 0, ["Sheet 'Raw Materials' not found"]

        ws = wb[sheet_name]
        count = 0
        errors = []

        # Read each column starting from column B
        col_num = 2
        while col_num <= ws.max_column:
            try:
                material_name = ws.cell(1, col_num).value
                if not material_name:
                    break

                frozen_rate_value = ws.cell(6, col_num).value
                frozen_rate = float(frozen_rate_value) if frozen_rate_value else None

                RawMaterial.objects.create(
                    quote=quote,
                    material_name=str(material_name),
                    grade=str(ws.cell(2, col_num).value or ''),
                    rm_code=str(ws.cell(3, col_num).value or ''),
                    unit_of_measurement=str(ws.cell(4, col_num).value or 'kg'),
                    rm_rate=float(ws.cell(5, col_num).value or 0),
                    frozen_rate=frozen_rate,
                    part_weight=float(ws.cell(7, col_num).value or 0),
                    runner_weight=float(ws.cell(8, col_num).value or 0),
                    process_losses=float(ws.cell(9, col_num).value or 0),
                    purging_loss_cost=float(ws.cell(10, col_num).value or 0),
                    other_rm_cost=float(ws.cell(11, col_num).value or 0),
                    other_rm_cost_description=str(ws.cell(12, col_num).value or ''),
                    icc_percentage=float(ws.cell(13, col_num).value or 0),
                    rejection_percentage=float(ws.cell(14, col_num).value or 0),
                    overhead_percentage=float(ws.cell(15, col_num).value or 0),
                    maintenance_percentage=float(ws.cell(16, col_num).value or 0),
                    profit_percentage=float(ws.cell(17, col_num).value or 0),
                )
                count += 1
                col_num += 1
            except Exception as e:
                errors.append(f"Column {get_column_letter(col_num)}: {str(e)}")
                col_num += 1

        return count, errors

    @staticmethod
    def parse_moulding_machines(file_path, quote):
        """Parse moulding machines from Excel (vertical format)"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        sheet_name = None
        for possible_name in ['Moulding Machines', 'Moulding_Machines', 'MouldingMachines']:
            if possible_name in wb.sheetnames:
                sheet_name = possible_name
                break

        if not sheet_name:
            return 0, ["Sheet 'Moulding Machines' not found"]

        ws = wb[sheet_name]
        count = 0
        errors = []

        col_num = 2
        while col_num <= ws.max_column:
            try:
                cavity = ws.cell(1, col_num).value
                if cavity is None:
                    break

                MouldingMachineDetail.objects.create(
                    quote=quote,
                    cavity=int(cavity),
                    machine_tonnage=float(ws.cell(2, col_num).value or 0),
                    cycle_time=float(ws.cell(3, col_num).value or 0),
                    efficiency=float(ws.cell(4, col_num).value or 0),
                    shift_rate=float(ws.cell(5, col_num).value or 0),
                    shift_rate_for_mtc=float(ws.cell(6, col_num).value or 0),
                    mtc_count=int(ws.cell(7, col_num).value or 0),
                    rejection_percentage=float(ws.cell(8, col_num).value or 0),
                    overhead_percentage=float(ws.cell(9, col_num).value or 0),
                    maintenance_percentage=float(ws.cell(10, col_num).value or 0),
                    profit_percentage=float(ws.cell(11, col_num).value or 0),
                )
                count += 1
                col_num += 1
            except Exception as e:
                errors.append(f"Column {get_column_letter(col_num)}: {str(e)}")
                col_num += 1

        return count, errors

    @staticmethod
    def parse_complete_quote(file_path, quote, user):
        """Parse a complete quote from Excel file and add components to existing quote"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        results = {
            'raw_materials': 0,
            'moulding_machines': 0,
            'assemblies': 0,
            'packaging': 0,
            'transport': 0,
        }
        errors = []

        # Raw Materials
        rm_sheet_name = 'Raw Materials' if 'Raw Materials' in wb.sheetnames else 'Raw_Materials'
        if rm_sheet_name in wb.sheetnames:
            ws_rm = wb[rm_sheet_name]
            col_num = 2
            while col_num <= ws_rm.max_column:
                try:
                    material_name = ws_rm.cell(1, col_num).value
                    if not material_name:
                        break

                    frozen_rate_value = ws_rm.cell(6, col_num).value
                    frozen_rate = float(frozen_rate_value) if frozen_rate_value else None

                    RawMaterial.objects.create(
                        quote=quote,
                        material_name=str(material_name),
                        grade=str(ws_rm.cell(2, col_num).value or ''),
                        rm_code=str(ws_rm.cell(3, col_num).value or ''),
                        unit_of_measurement=str(ws_rm.cell(4, col_num).value or 'kg'),
                        rm_rate=float(ws_rm.cell(5, col_num).value or 0),
                        frozen_rate=frozen_rate,
                        part_weight=float(ws_rm.cell(7, col_num).value or 0),
                        runner_weight=float(ws_rm.cell(8, col_num).value or 0),
                        process_losses=float(ws_rm.cell(9, col_num).value or 0),
                        purging_loss_cost=float(ws_rm.cell(10, col_num).value or 0),
                        other_rm_cost=float(ws_rm.cell(11, col_num).value or 0),
                        other_rm_cost_description=str(ws_rm.cell(12, col_num).value or ''),
                        icc_percentage=float(ws_rm.cell(13, col_num).value or 0),
                        rejection_percentage=float(ws_rm.cell(14, col_num).value or 0),
                        overhead_percentage=float(ws_rm.cell(15, col_num).value or 0),
                        maintenance_percentage=float(ws_rm.cell(16, col_num).value or 0),
                        profit_percentage=float(ws_rm.cell(17, col_num).value or 0),
                    )
                    results['raw_materials'] += 1
                    col_num += 1
                except Exception as e:
                    errors.append(f"Raw Material Column {get_column_letter(col_num)}: {str(e)}")
                    col_num += 1

        # Moulding Machines
        mm_sheet_name = 'Moulding Machines' if 'Moulding Machines' in wb.sheetnames else 'Moulding_Machines'
        if mm_sheet_name in wb.sheetnames:
            ws_mm = wb[mm_sheet_name]
            col_num = 2
            while col_num <= ws_mm.max_column:
                try:
                    cavity = ws_mm.cell(1, col_num).value
                    if cavity is None:
                        break

                    MouldingMachineDetail.objects.create(
                        quote=quote,
                        cavity=int(cavity),
                        machine_tonnage=float(ws_mm.cell(2, col_num).value or 0),
                        cycle_time=float(ws_mm.cell(3, col_num).value or 0),
                        efficiency=float(ws_mm.cell(4, col_num).value or 0),
                        shift_rate=float(ws_mm.cell(5, col_num).value or 0),
                        shift_rate_for_mtc=float(ws_mm.cell(6, col_num).value or 0),
                        mtc_count=int(ws_mm.cell(7, col_num).value or 0),
                        rejection_percentage=float(ws_mm.cell(8, col_num).value or 0),
                        overhead_percentage=float(ws_mm.cell(9, col_num).value or 0),
                        maintenance_percentage=float(ws_mm.cell(10, col_num).value or 0),
                        profit_percentage=float(ws_mm.cell(11, col_num).value or 0),
                    )
                    results['moulding_machines'] += 1
                    col_num += 1
                except Exception as e:
                    errors.append(f"Moulding Machine Column {get_column_letter(col_num)}: {str(e)}")
                    col_num += 1

        # Assemblies
        asm_sheet_name = 'Assemblies' if 'Assemblies' in wb.sheetnames else 'Assemblies'
        if asm_sheet_name in wb.sheetnames:
            ws_asm = wb[asm_sheet_name]
            col_num = 2
            while col_num <= ws_asm.max_column:
                try:
                    assembly_name = ws_asm.cell(1, col_num).value
                    if not assembly_name:
                        break

                    Assembly.objects.create(
                        quote=quote,
                        name=str(assembly_name),
                        remarks=str(ws_asm.cell(3, col_num).value or ''),
                        manual_cost=float(ws_asm.cell(4, col_num).value or 0),
                        other_cost=float(ws_asm.cell(5, col_num).value or 0),
                        other_cost_description=str(ws_asm.cell(6, col_num).value or ''),
                        inspection_handling_cost=float(ws_asm.cell(7, col_num).value or 0),
                        profit_percentage=float(ws_asm.cell(8, col_num).value or 0),
                        rejection_percentage=float(ws_asm.cell(9, col_num).value or 0),
                    )
                    results['assemblies'] += 1
                    col_num += 1
                except Exception as e:
                    errors.append(f"Assembly Column {get_column_letter(col_num)}: {str(e)}")
                    col_num += 1

        # Packaging
        pkg_sheet_name = 'Packaging' if 'Packaging' in wb.sheetnames else 'Packaging'
        if pkg_sheet_name in wb.sheetnames:
            ws_pkg = wb[pkg_sheet_name]
            col_num = 2
            while col_num <= ws_pkg.max_column:
                try:
                    category = ws_pkg.cell(1, col_num).value
                    if not category:
                        break

                    Packaging.objects.create(
                        quote=quote,
                        packaging_type=None,
                        packaging_category=str(category or 'box'),
                        parts_per_packaging=int(ws_pkg.cell(2, col_num).value or 0),
                        maintenance_percentage=float(ws_pkg.cell(3, col_num).value or 0),
                        # Box fields (rows 5-9)
                        packaging_length=float(ws_pkg.cell(5, col_num).value or 0),
                        packaging_breadth=float(ws_pkg.cell(6, col_num).value or 0),
                        packaging_height=float(ws_pkg.cell(7, col_num).value or 0),
                        cost=float(ws_pkg.cell(8, col_num).value or 0),
                        lifecycle=int(ws_pkg.cell(9, col_num).value or 0),
                        # Polybag fields (rows 11-14)
                        polybag_length=float(ws_pkg.cell(11, col_num).value or 0),
                        polybag_width=float(ws_pkg.cell(12, col_num).value or 0),
                        rate_per_kg=float(ws_pkg.cell(13, col_num).value or 0),
                        polybags_per_kg=float(ws_pkg.cell(14, col_num).value or 0),
                    )
                    results['packaging'] += 1
                    col_num += 1
                except Exception as e:
                    errors.append(f"Packaging Column {get_column_letter(col_num)}: {str(e)}")
                    col_num += 1

        # Transport
        trans_sheet_name = 'Transport' if 'Transport' in wb.sheetnames else 'Transport'
        if trans_sheet_name in wb.sheetnames:
            ws_trans = wb[trans_sheet_name]
            col_num = 2
            while col_num <= ws_trans.max_column:
                try:
                    length = ws_trans.cell(1, col_num).value
                    if length is None:
                        break

                    Transport.objects.create(
                        quote=quote,
                        transport_length=float(length or 0),
                        transport_breadth=float(ws_trans.cell(2, col_num).value or 0),
                        transport_height=float(ws_trans.cell(3, col_num).value or 0),
                        trip_cost=float(ws_trans.cell(4, col_num).value or 0),
                        parts_per_box=int(ws_trans.cell(5, col_num).value or 0),
                    )
                    results['transport'] += 1
                    col_num += 1
                except Exception as e:
                    errors.append(f"Transport Column {get_column_letter(col_num)}: {str(e)}")
                    col_num += 1

        return quote, errors, results

    @staticmethod
    def parse_multiple_quotes_complete(file_path, project, customer_group, user):
        """Parse multiple complete quotes from horizontal format Excel file"""
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            for chunk in file_path.chunks():
                tmp_file.write(chunk)
            tmp_file_path = tmp_file.name

        wb = openpyxl.load_workbook(tmp_file_path, data_only=True)

        results = {
            'quotes': 0,
            'components': {
                'raw_materials': 0,
                'moulding_machines': 0,
                'assemblies': 0,
                'packaging': 0,
                'transport': 0,
            },
            'errors': []
        }

        try:
            # Parse Quote Definition sheet
            if 'Quote Definition' not in wb.sheetnames:
                results['errors'].append("Quote Definition sheet not found")
                return results

            ws_def = wb['Quote Definition']

            # Read each column starting from column B
            col_num = 2
            while col_num <= ws_def.max_column:
                try:
                    quote_name = ws_def.cell(1, col_num).value
                    if not quote_name:
                        break

                    # Create quote
                    quote = Quote.objects.create(
                        project=project,
                        name=str(quote_name),
                        client_group=customer_group,
                        client_name=str(ws_def.cell(2, col_num).value or ''),
                        sap_number=str(ws_def.cell(3, col_num).value or ''),
                        part_number=str(ws_def.cell(4, col_num).value or ''),
                        part_name=str(ws_def.cell(5, col_num).value or ''),
                        amendment_number=str(ws_def.cell(6, col_num).value or ''),
                        description=str(ws_def.cell(7, col_num).value or ''),
                        quantity=int(ws_def.cell(8, col_num).value or 1),
                        handling_charge=float(ws_def.cell(9, col_num).value or 0),
                        profit_percentage=float(ws_def.cell(10, col_num).value or 0),
                        notes=str(ws_def.cell(11, col_num).value or ''),
                        created_by=user,
                        quote_definition_complete=True
                    )

                    # Parse components for this quote
                    component_errors = ExcelParser._parse_components_horizontal(wb, quote, str(quote_name))
                    results['errors'].extend(component_errors)

                    results['quotes'] += 1
                    col_num += 1

                except Exception as e:
                    results['errors'].append(f"Quote Column {get_column_letter(col_num)}: {str(e)}")
                    col_num += 1

            # Count total components across all quotes
            for quote in Quote.objects.filter(project=project, created_by=user):
                results['components']['raw_materials'] += quote.raw_materials.count()
                results['components']['moulding_machines'] += quote.moulding_machines.count()
                results['components']['assemblies'] += quote.assemblies.count()
                results['components']['packaging'] += quote.packagings.count()
                results['components']['transport'] += quote.transports.count()

        finally:
            # Clean up temp file
            os.unlink(tmp_file_path)

        return results

    @staticmethod
    def _parse_components_horizontal(wb, quote, quote_name):
        """Parse components for a single quote from horizontal format sheets"""
        errors = []

        # Parse Raw Materials
        rm_sheet_name = 'Raw Materials' if 'Raw Materials' in wb.sheetnames else 'Raw_Materials'
        if rm_sheet_name in wb.sheetnames:
            ws_rm = wb[rm_sheet_name]
            col_num = 2
            while col_num <= ws_rm.max_column:
                try:
                    # Check if this column belongs to our quote
                    col_quote_name = ws_rm.cell(row=1, column=col_num).value
                    if not col_quote_name or str(col_quote_name) != quote_name:
                        col_num += 1
                        continue

                    material_name = ws_rm.cell(row=2, column=col_num).value
                    if not material_name:
                        col_num += 1
                        continue

                    # Get frozen rate (might be None)
                    frozen_rate_value = ws_rm.cell(row=7, column=col_num).value
                    frozen_rate = float(frozen_rate_value) if frozen_rate_value else None

                    RawMaterial.objects.create(
                        quote=quote,
                        material_name=str(material_name),
                        grade=str(ws_rm.cell(row=3, column=col_num).value or ''),
                        rm_code=str(ws_rm.cell(row=4, column=col_num).value or ''),
                        unit_of_measurement=str(ws_rm.cell(row=5, column=col_num).value or 'kg'),
                        rm_rate=float(ws_rm.cell(row=6, column=col_num).value or 0),
                        frozen_rate=frozen_rate,
                        part_weight=float(ws_rm.cell(row=8, column=col_num).value or 0),
                        runner_weight=float(ws_rm.cell(row=9, column=col_num).value or 0),
                        process_losses=float(ws_rm.cell(row=10, column=col_num).value or 0),
                        purging_loss_cost=float(ws_rm.cell(row=11, column=col_num).value or 0),
                        other_rm_cost=float(ws_rm.cell(row=12, column=col_num).value or 0),
                        other_rm_cost_description=str(ws_rm.cell(row=13, column=col_num).value or ''),
                        icc_percentage=float(ws_rm.cell(row=14, column=col_num).value or 0),
                        rejection_percentage=float(ws_rm.cell(row=15, column=col_num).value or 0),
                        overhead_percentage=float(ws_rm.cell(row=16, column=col_num).value or 0),
                        maintenance_percentage=float(ws_rm.cell(row=17, column=col_num).value or 0),
                        profit_percentage=float(ws_rm.cell(row=18, column=col_num).value or 0),
                    )
                    col_num += 1
                except Exception as e:
                    errors.append(f"Raw Material - {quote_name} Column {get_column_letter(col_num)}: {str(e)}")
                    col_num += 1

        # Parse Moulding Machines
        mm_sheet_name = 'Moulding Machines' if 'Moulding Machines' in wb.sheetnames else 'Moulding_Machines'
        if mm_sheet_name in wb.sheetnames:
            ws_mm = wb[mm_sheet_name]
            col_num = 2
            while col_num <= ws_mm.max_column:
                try:
                    # Check if this column belongs to our quote
                    col_quote_name = ws_mm.cell(row=1, column=col_num).value
                    if not col_quote_name or str(col_quote_name) != quote_name:
                        col_num += 1
                        continue

                    cavity = ws_mm.cell(row=2, column=col_num).value
                    if cavity is None:
                        col_num += 1
                        continue

                    MouldingMachineDetail.objects.create(
                        quote=quote,
                        cavity=int(cavity),
                        machine_tonnage=float(ws_mm.cell(row=3, column=col_num).value or 0),
                        cycle_time=float(ws_mm.cell(row=4, column=col_num).value or 0),
                        efficiency=float(ws_mm.cell(row=5, column=col_num).value or 0),
                        shift_rate=float(ws_mm.cell(row=6, column=col_num).value or 0),
                        shift_rate_for_mtc=float(ws_mm.cell(row=7, column=col_num).value or 0),
                        mtc_count=int(ws_mm.cell(row=8, column=col_num).value or 0),
                        rejection_percentage=float(ws_mm.cell(row=9, column=col_num).value or 0),
                        overhead_percentage=float(ws_mm.cell(row=10, column=col_num).value or 0),
                        maintenance_percentage=float(ws_mm.cell(row=11, column=col_num).value or 0),
                        profit_percentage=float(ws_mm.cell(row=12, column=col_num).value or 0),
                    )
                    col_num += 1
                except Exception as e:
                    errors.append(f"Moulding Machine - {quote_name} Column {get_column_letter(col_num)}: {str(e)}")
                    col_num += 1

        # Parse Assemblies
        asm_sheet_name = 'Assemblies' if 'Assemblies' in wb.sheetnames else 'Assemblies'
        if asm_sheet_name in wb.sheetnames:
            ws_asm = wb[asm_sheet_name]
            col_num = 2
            while col_num <= ws_asm.max_column:
                try:
                    # Check if this column belongs to our quote
                    col_quote_name = ws_asm.cell(row=1, column=col_num).value
                    if not col_quote_name or str(col_quote_name) != quote_name:
                        col_num += 1
                        continue

                    assembly_name = ws_asm.cell(row=2, column=col_num).value
                    if not assembly_name:
                        col_num += 1
                        continue

                    Assembly.objects.create(
                        quote=quote,
                        name=str(assembly_name),
                        remarks=str(ws_asm.cell(row=4, column=col_num).value or ''),
                        manual_cost=float(ws_asm.cell(row=5, column=col_num).value or 0),
                        other_cost=float(ws_asm.cell(row=6, column=col_num).value or 0),
                        other_cost_description=str(ws_asm.cell(row=7, column=col_num).value or ''),
                        inspection_handling_cost=float(ws_asm.cell(row=8, column=col_num).value or 0),
                        profit_percentage=float(ws_asm.cell(row=9, column=col_num).value or 0),
                        rejection_percentage=float(ws_asm.cell(row=10, column=col_num).value or 0),
                    )
                    col_num += 1
                except Exception as e:
                    errors.append(f"Assembly - {quote_name} Column {get_column_letter(col_num)}: {str(e)}")
                    col_num += 1

        # Parse Packaging
        pkg_sheet_name = 'Packaging' if 'Packaging' in wb.sheetnames else 'Packaging'
        if pkg_sheet_name in wb.sheetnames:
            ws_pkg = wb[pkg_sheet_name]
            col_num = 2
            while col_num <= ws_pkg.max_column:
                try:
                    # Check if this column belongs to our quote
                    col_quote_name = ws_pkg.cell(row=1, column=col_num).value
                    if not col_quote_name or str(col_quote_name) != quote_name:
                        col_num += 1
                        continue

                    category = ws_pkg.cell(row=2, column=col_num).value
                    if not category:
                        col_num += 1
                        continue

                    Packaging.objects.create(
                        quote=quote,
                        packaging_type=None,
                        packaging_category=str(category or 'box'),
                        parts_per_packaging=int(ws_pkg.cell(row=3, column=col_num).value or 0),
                        maintenance_percentage=float(ws_pkg.cell(row=4, column=col_num).value or 0),
                        # Box fields
                        packaging_length=float(ws_pkg.cell(row=6, column=col_num).value or 0),
                        packaging_breadth=float(ws_pkg.cell(row=7, column=col_num).value or 0),
                        packaging_height=float(ws_pkg.cell(row=8, column=col_num).value or 0),
                        cost=float(ws_pkg.cell(row=9, column=col_num).value or 0),
                        lifecycle=int(ws_pkg.cell(row=10, column=col_num).value or 0),
                        # Polybag fields
                        polybag_length=float(ws_pkg.cell(row=12, column=col_num).value or 0),
                        polybag_width=float(ws_pkg.cell(row=13, column=col_num).value or 0),
                        rate_per_kg=float(ws_pkg.cell(row=14, column=col_num).value or 0),
                        polybags_per_kg=float(ws_pkg.cell(row=15, column=col_num).value or 0),
                    )
                    col_num += 1
                except Exception as e:
                    errors.append(f"Packaging - {quote_name} Column {get_column_letter(col_num)}: {str(e)}")
                    col_num += 1

        # Parse Transport
        trans_sheet_name = 'Transport' if 'Transport' in wb.sheetnames else 'Transport'
        if trans_sheet_name in wb.sheetnames:
            ws_trans = wb[trans_sheet_name]
            col_num = 2
            while col_num <= ws_trans.max_column:
                try:
                    # Check if this column belongs to our quote
                    col_quote_name = ws_trans.cell(row=1, column=col_num).value
                    if not col_quote_name or str(col_quote_name) != quote_name:
                        col_num += 1
                        continue

                    length = ws_trans.cell(row=2, column=col_num).value
                    if length is None:
                        col_num += 1
                        continue

                    Transport.objects.create(
                        quote=quote,
                        transport_length=float(length or 0),
                        transport_breadth=float(ws_trans.cell(row=3, column=col_num).value or 0),
                        transport_height=float(ws_trans.cell(row=4, column=col_num).value or 0),
                        trip_cost=float(ws_trans.cell(row=5, column=col_num).value or 0),
                        parts_per_box=int(ws_trans.cell(row=6, column=col_num).value or 0),
                    )
                    col_num += 1
                except Exception as e:
                    errors.append(f"Transport - {quote_name} Column {get_column_letter(col_num)}: {str(e)}")
                    col_num += 1

        return errors


class ExcelExporter:
    """Export quotes to Excel with all calculated fields"""

    @staticmethod
    def _add_vertical_headers(ws, headers, color, start_row=1):
        """Add vertical headers in column A with color coding"""
        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for row_num, header in enumerate(headers, start_row):
            cell = ws.cell(row=row_num, column=1)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.column_dimensions['A'].width = 35

    @staticmethod
    def export_quote(quote):
        """Export a single quote with all components and calculated fields"""
        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: Quote Definition
        ws_def = wb.create_sheet("Quote Definition")
        def_headers = [
            'Field',
            'Value'
        ]

        # Add headers
        ws_def.cell(row=1, column=1, value='Field').font = Font(bold=True)
        ws_def.cell(row=1, column=2, value='Value').font = Font(bold=True)

        # Add data
        quote_data = [
            ('Quote Name', quote.name),
            ('Version', f"{quote.major_version}.{quote.minor_version}"),
            ('Client Group', quote.client_group.name if quote.client_group else ''),
            ('Client Name', quote.client_name),
            ('SAP Number', quote.sap_number),
            ('Part Number', quote.part_number),
            ('Part Name', quote.part_name),
            ('Amendment Number', quote.amendment_number),
            ('Description', quote.description),
            ('Quantity', quote.quantity),
            ('Handling Charge', float(quote.handling_charge)),
            ('Profit %', float(quote.profit_percentage)),
            ('Status', quote.get_status_display()),
            ('Notes', quote.notes),
        ]

        for row_num, (field, value) in enumerate(quote_data, 2):
            ws_def.cell(row=row_num, column=1, value=field)
            ws_def.cell(row=row_num, column=2, value=value)

        ws_def.column_dimensions['A'].width = 25
        ws_def.column_dimensions['B'].width = 50

        # Sheet 2: Raw Materials (vertical format with calculated fields)
        if quote.raw_materials.exists():
            ws_rm = wb.create_sheet("Raw Materials")
            rm_headers = [
                'Material Name',
                'Grade',
                'RM Code',
                'Unit',
                'RM Rate (per kg)',
                'Frozen Rate (per kg)',
                'Effective Rate (per kg)',
                'Part Weight',
                'Runner Weight',
                'Process Losses %',
                'Purging Losses %',
                'Gross Weight (grams)',
                'Other RM Cost',
                'Other RM Cost Description',
                'ICC %',
                'Rejection %',
                'Overhead %',
                'Maintenance %',
                'Profit %',
                'Base RM Cost',
                'Frozen RM Cost',
                'Total (Without Profit)',
                'Total Cost'
            ]

            ExcelExporter._add_vertical_headers(ws_rm, rm_headers, "4472C4")

            for col_num, rm in enumerate(quote.raw_materials.all(), 2):
                ws_rm.cell(row=1, column=col_num, value=rm.material_name)
                ws_rm.cell(row=2, column=col_num, value=rm.grade)
                ws_rm.cell(row=3, column=col_num, value=rm.rm_code)
                ws_rm.cell(row=4, column=col_num, value=rm.unit_of_measurement)
                ws_rm.cell(row=5, column=col_num, value=float(rm.rm_rate))
                ws_rm.cell(row=6, column=col_num, value=float(rm.frozen_rate) if rm.frozen_rate else None)
                ws_rm.cell(row=7, column=col_num, value=float(rm.effective_rate_per_kg))
                ws_rm.cell(row=8, column=col_num, value=float(rm.part_weight))
                ws_rm.cell(row=9, column=col_num, value=float(rm.runner_weight))
                ws_rm.cell(row=10, column=col_num, value=float(rm.process_losses))
                ws_rm.cell(row=11, column=col_num, value=float(rm.purging_loss_cost))
                ws_rm.cell(row=12, column=col_num, value=float(rm.gross_weight_in_grams))
                ws_rm.cell(row=13, column=col_num, value=float(rm.other_rm_cost))
                ws_rm.cell(row=14, column=col_num, value=rm.other_rm_cost_description)
                ws_rm.cell(row=15, column=col_num, value=float(rm.icc_percentage))
                ws_rm.cell(row=16, column=col_num, value=float(rm.rejection_percentage))
                ws_rm.cell(row=17, column=col_num, value=float(rm.overhead_percentage))
                ws_rm.cell(row=18, column=col_num, value=float(rm.maintenance_percentage))
                ws_rm.cell(row=19, column=col_num, value=float(rm.profit_percentage))
                ws_rm.cell(row=20, column=col_num, value=float(rm.base_rm_cost))
                ws_rm.cell(row=21, column=col_num, value=float(rm.frozen_rm_cost) if rm.frozen_rm_cost else None)
                ws_rm.cell(row=22, column=col_num, value=float(rm.total_rm_cost_without_profit))
                ws_rm.cell(row=23, column=col_num, value=float(rm.rm_cost))
                ws_rm.column_dimensions[get_column_letter(col_num)].width = 18

        # Sheet 3: Moulding Machines (vertical format with calculated fields)
        if quote.moulding_machines.exists():
            ws_mm = wb.create_sheet("Moulding Machines")
            mm_headers = [
                'Cavity',
                'Machine Tonnage',
                'Cycle Time (s)',
                'Efficiency %',
                'Shift Rate',
                'Shift Rate for MTC',
                'MTC Count',
                'Rejection %',
                'Overhead %',
                'Maintenance %',
                'Profit %',
                'Parts per Shift',
                'MTC Cost',
                'Base Conversion Cost',
                'Rejection Cost',
                'Overhead Cost',
                'Maintenance Cost',
                'Profit Cost',
                'Total Conversion Cost'
            ]

            ExcelExporter._add_vertical_headers(ws_mm, mm_headers, "70AD47")

            for col_num, mm in enumerate(quote.moulding_machines.all(), 2):
                ws_mm.cell(row=1, column=col_num, value=mm.cavity)
                ws_mm.cell(row=2, column=col_num, value=float(mm.machine_tonnage))
                ws_mm.cell(row=3, column=col_num, value=float(mm.cycle_time))
                ws_mm.cell(row=4, column=col_num, value=float(mm.efficiency))
                ws_mm.cell(row=5, column=col_num, value=float(mm.shift_rate))
                ws_mm.cell(row=6, column=col_num, value=float(mm.shift_rate_for_mtc))
                ws_mm.cell(row=7, column=col_num, value=mm.mtc_count)
                ws_mm.cell(row=8, column=col_num, value=float(mm.rejection_percentage))
                ws_mm.cell(row=9, column=col_num, value=float(mm.overhead_percentage))
                ws_mm.cell(row=10, column=col_num, value=float(mm.maintenance_percentage))
                ws_mm.cell(row=11, column=col_num, value=float(mm.profit_percentage))
                ws_mm.cell(row=12, column=col_num, value=mm.number_of_parts_per_shift)
                ws_mm.cell(row=13, column=col_num, value=float(mm.mtc_cost))
                ws_mm.cell(row=14, column=col_num, value=float(mm.base_conversion_cost))
                ws_mm.cell(row=15, column=col_num, value=float(mm.rejection_cost))
                ws_mm.cell(row=16, column=col_num, value=float(mm.overhead_cost))
                ws_mm.cell(row=17, column=col_num, value=float(mm.machine_maintenance_cost))
                ws_mm.cell(row=18, column=col_num, value=float(mm.machine_profit_cost))
                ws_mm.cell(row=19, column=col_num, value=float(mm.conversion_cost))
                ws_mm.column_dimensions[get_column_letter(col_num)].width = 18

        # Sheet 4: Assemblies (vertical format with calculated fields)
        if quote.assemblies.exists():
            ws_asm = wb.create_sheet("Assemblies")
            asm_headers = [
                'Assembly Name',
                'Remarks',
                'Manual Cost',
                'Other Cost',
                'Other Cost Description',
                'Inspection & Handling Cost',
                'Profit %',
                'Rejection %',
                'Base Cost',
                'Profit Cost',
                'Rejection Cost',
                'Total Cost'
            ]

            ExcelExporter._add_vertical_headers(ws_asm, asm_headers, "FFC000")

            for col_num, asm in enumerate(quote.assemblies.all(), 2):
                costs = asm.calculate_costs()
                ws_asm.cell(row=1, column=col_num, value=asm.name)
                ws_asm.cell(row=2, column=col_num, value=asm.remarks)
                ws_asm.cell(row=3, column=col_num, value=float(asm.manual_cost))
                ws_asm.cell(row=4, column=col_num, value=float(asm.other_cost))
                ws_asm.cell(row=5, column=col_num, value=asm.other_cost_description)
                ws_asm.cell(row=6, column=col_num, value=float(asm.inspection_handling_cost))
                ws_asm.cell(row=7, column=col_num, value=float(asm.profit_percentage))
                ws_asm.cell(row=8, column=col_num, value=float(asm.rejection_percentage))
                ws_asm.cell(row=9, column=col_num, value=float(costs['base_cost']))
                ws_asm.cell(row=10, column=col_num, value=float(costs['profit_cost']))
                ws_asm.cell(row=11, column=col_num, value=float(costs['rejection_cost']))
                ws_asm.cell(row=12, column=col_num, value=float(costs['total_assembly_cost']))
                ws_asm.column_dimensions[get_column_letter(col_num)].width = 18

        # Sheet 5: Packaging (vertical format with calculated fields)
        if quote.packagings.exists():
            ws_pkg = wb.create_sheet("Packaging")
            pkg_headers = [
                'Packaging Category',
                'Packaging Type',
                'Parts per Packaging',
                'Maintenance %',
                '--- BOX FIELDS ---',
                'Length (mm)',
                'Breadth (mm)',
                'Height (mm)',
                'Cost',
                'Lifecycle',
                '--- POLYBAG FIELDS ---',
                'Polybag Length (inches)',
                'Polybag Width (inches)',
                'Rate per kg',
                'Polybags per kg',
                '--- CALCULATED ---',
                'Maintenance Cost',
                'Cost per Part',
                'Total Cost'
            ]

            ExcelExporter._add_vertical_headers(ws_pkg, pkg_headers, "E26B0A")

            for col_num, pkg in enumerate(quote.packagings.all(), 2):
                ws_pkg.cell(row=1, column=col_num, value=pkg.get_packaging_category_display())
                ws_pkg.cell(row=2, column=col_num, value=pkg.packaging_type.name if pkg.packaging_type else 'Custom')
                ws_pkg.cell(row=3, column=col_num, value=pkg.parts_per_packaging)
                ws_pkg.cell(row=4, column=col_num, value=float(pkg.maintenance_percentage))
                # Box fields
                ws_pkg.cell(row=6, column=col_num, value=float(pkg.packaging_length))
                ws_pkg.cell(row=7, column=col_num, value=float(pkg.packaging_breadth))
                ws_pkg.cell(row=8, column=col_num, value=float(pkg.packaging_height))
                ws_pkg.cell(row=9, column=col_num, value=float(pkg.cost))
                ws_pkg.cell(row=10, column=col_num, value=pkg.lifecycle)
                # Polybag fields
                ws_pkg.cell(row=12, column=col_num, value=float(pkg.polybag_length))
                ws_pkg.cell(row=13, column=col_num, value=float(pkg.polybag_width))
                ws_pkg.cell(row=14, column=col_num, value=float(pkg.rate_per_kg))
                ws_pkg.cell(row=15, column=col_num, value=float(pkg.polybags_per_kg))
                # Calculated
                ws_pkg.cell(row=17, column=col_num, value=float(pkg.maintenance_cost))
                ws_pkg.cell(row=18, column=col_num, value=float(pkg.cost_per_part))
                ws_pkg.cell(row=19, column=col_num, value=float(pkg.total_cost))
                ws_pkg.column_dimensions[get_column_letter(col_num)].width = 18

        # Sheet 6: Transport (vertical format with calculated fields)
        if quote.transports.exists():
            ws_trans = wb.create_sheet("Transport")
            trans_headers = [
                'Length (ft)',
                'Breadth (ft)',
                'Height (ft)',
                'Trip Cost',
                'Parts per Box',
                'Length (mm)',
                'Breadth (mm)',
                'Height (mm)',
                'Boxes on Length',
                'Boxes on Breadth',
                'Boxes on Height',
                'Total Boxes',
                'Total Parts per Trip',
                'Trip Cost per Part'
            ]

            ExcelExporter._add_vertical_headers(ws_trans, trans_headers, "9933FF")

            for col_num, trans in enumerate(quote.transports.all(), 2):
                ws_trans.cell(row=1, column=col_num, value=float(trans.transport_length))
                ws_trans.cell(row=2, column=col_num, value=float(trans.transport_breadth))
                ws_trans.cell(row=3, column=col_num, value=float(trans.transport_height))
                ws_trans.cell(row=4, column=col_num, value=float(trans.trip_cost))
                ws_trans.cell(row=5, column=col_num, value=trans.parts_per_box)
                ws_trans.cell(row=6, column=col_num, value=float(trans.transport_length_mm))
                ws_trans.cell(row=7, column=col_num, value=float(trans.transport_breadth_mm))
                ws_trans.cell(row=8, column=col_num, value=float(trans.transport_height_mm))
                ws_trans.cell(row=9, column=col_num, value=trans.boxes_on_length)
                ws_trans.cell(row=10, column=col_num, value=trans.boxes_on_breadth)
                ws_trans.cell(row=11, column=col_num, value=trans.boxes_on_height)
                ws_trans.cell(row=12, column=col_num, value=trans.total_boxes)
                ws_trans.cell(row=13, column=col_num, value=trans.total_parts_per_trip)
                ws_trans.cell(row=14, column=col_num, value=float(trans.trip_cost_per_part))
                ws_trans.column_dimensions[get_column_letter(col_num)].width = 18

        # Sheet 7: Summary
        ws_summary = wb.create_sheet("Summary", 0)  # Insert at beginning
        summary_headers = [
            'Description',
            'Amount'
        ]

        ws_summary.cell(row=1, column=1, value='Description').font = Font(bold=True)
        ws_summary.cell(row=1, column=2, value='Amount').font = Font(bold=True)

        summary_data = [
            ('Quote Name', quote.name),
            ('Version', f"{quote.major_version}.{quote.minor_version}"),
            ('Quantity', quote.quantity),
            ('', ''),
            ('Total Raw Material Cost', float(quote.get_total_raw_material_cost())),
            ('Total Conversion Cost', float(quote.get_total_conversion_cost())),
            ('Total Assembly Cost', float(quote.get_total_assembly_cost())),
            ('Total Packaging Cost', float(quote.get_total_packaging_cost())),
            ('Total Transport Cost', float(quote.get_total_transport_cost())),
            ('', ''),
            ('Base Cost', float(quote.get_base_cost())),
            ('Profit Amount', float(quote.get_profit_amount())),
            ('Handling Charge', float(quote.handling_charge)),
            ('', ''),
            ('Grand Total', float(quote.get_grand_total())),
            ('Cost per Part', float(quote.get_grand_total()) / quote.quantity if quote.quantity > 0 else 0),
        ]

        for row_num, (desc, amt) in enumerate(summary_data, 2):
            ws_summary.cell(row=row_num, column=1, value=desc)
            if isinstance(amt, (int, float)):
                ws_summary.cell(row=row_num, column=2, value=amt)

        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20

        return wb

    @staticmethod
    def export_project(project):
        """Export entire project with all quotes"""
        wb = Workbook()
        wb.remove(wb.active)

        # Create a summary sheet
        ws_summary = wb.create_sheet("Project Summary")
        ws_summary.cell(row=1, column=1, value='Project Name').font = Font(bold=True)
        ws_summary.cell(row=1, column=2, value=project.name)
        ws_summary.cell(row=2, column=1, value='Description').font = Font(bold=True)
        ws_summary.cell(row=2, column=2, value=project.description)
        ws_summary.cell(row=3, column=1, value='Total Quotes').font = Font(bold=True)
        ws_summary.cell(row=3, column=2, value=project.quotes.count())

        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 50

        # Add a sheet for each quote
        for quote in project.quotes.all():
            # Export each quote to its own workbook
            quote_wb = ExcelExporter.export_quote(quote)

            # Copy all sheets from quote workbook to project workbook
            for sheet_name in quote_wb.sheetnames:
                source_sheet = quote_wb[sheet_name]
                # Create unique sheet name with quote name prefix
                target_sheet_name = f"{quote.name[:20]}_{sheet_name}"[:31]  # Excel sheet name limit
                target_sheet = wb.create_sheet(target_sheet_name)

                # Copy all cells
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_cell = target_sheet[cell.coordinate]
                        target_cell.value = cell.value
                        if cell.has_style:
                            target_cell.font = cell.font.copy()
                            target_cell.fill = cell.fill.copy()
                            target_cell.alignment = cell.alignment.copy()

                # Copy column dimensions
                for col_letter, col_dim in source_sheet.column_dimensions.items():
                    target_sheet.column_dimensions[col_letter].width = col_dim.width

        return wb