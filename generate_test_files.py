"""
Script to generate test Excel files for quote creation
Run this script to create sample Excel files in a 'test_files' directory
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_test_files_directory():
    """Create test_files directory if it doesn't exist"""
    if not os.path.exists('test_files'):
        os.makedirs('test_files')
    print("✓ Created test_files directory")


def add_vertical_headers(worksheet, headers, color):
    """Helper to add vertical headers to worksheet"""
    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for row_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=1)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    worksheet.column_dimensions['A'].width = 30


def generate_raw_materials_test():
    """Generate test file for raw materials upload"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Materials"
    
    headers = [
        'Material Name*', 'Grade', 'RM Code*', 'Unit (kg/gm/ton)*',
        'RM Rate*', 'Frozen Rate', 'Part Weight*', 'Runner Weight*',
        'Process Losses', 'Purging Loss Cost', 'ICC %',
        'Rejection %', 'Overhead %', 'Maintenance %', 'Profit %'
    ]
    
    add_vertical_headers(ws, headers, "4472C4")
    
    # Add 5 sample raw materials
    materials = [
        ['Polypropylene', 'PP-H340R', 'RM-PP-001', 'kg', 125.50, None, 0.0234, 0.0045, 2.5, 1.2, 0.5, 2.0, 5.0, 3.0, 10.0],
        ['ABS Resin', 'ABS-750', 'RM-ABS-002', 'kg', 185.75, 180.00, 0.0456, 0.0089, 3.0, 1.5, 0.75, 2.5, 4.5, 2.5, 12.0],
        ['Nylon 6', 'PA6-GF30', 'RM-PA6-003', 'kg', 215.00, None, 0.0678, 0.0123, 4.5, 2.0, 1.0, 3.0, 6.0, 4.0, 15.0],
        ['Polycarbonate', 'PC-1500', 'RM-PC-004', 'kg', 295.50, 290.00, 0.0345, 0.0067, 2.0, 1.0, 0.5, 1.5, 4.0, 2.0, 8.0],
        ['PVC Compound', 'PVC-SG5', 'RM-PVC-005', 'kg', 95.25, None, 0.0289, 0.0056, 1.5, 0.8, 0.25, 1.8, 3.5, 2.2, 9.0],
    ]
    
    for col_num, material in enumerate(materials, 2):
        for row_num, value in enumerate(material, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    wb.save('test_files/test_raw_materials.xlsx')
    print("✓ Generated test_raw_materials.xlsx (5 raw materials)")


def generate_moulding_machines_test():
    """Generate test file for moulding machines upload"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Moulding Machines"
    
    headers = [
        'Cavity*', 'Machine Tonnage*', 'Cycle Time (s)*', 'Efficiency %*',
        'Shift Rate*', 'Shift Rate for MTC*', 'MTC Count*',
        'Rejection %', 'Overhead %', 'Maintenance %', 'Profit %'
    ]
    
    add_vertical_headers(ws, headers, "70AD47")
    
    # Add 4 sample machines
    machines = [
        [2, 120, 38.5, 87.5, 4500.00, 4200.00, 2, 2.5, 5.0, 3.5, 15.0],
        [4, 180, 45.2, 90.0, 6000.00, 5500.00, 3, 2.0, 4.5, 3.0, 12.0],
        [8, 250, 52.8, 85.0, 8500.00, 8000.00, 4, 3.0, 6.0, 4.0, 18.0],
        [1, 80, 28.3, 92.0, 3200.00, 3000.00, 1, 1.5, 4.0, 2.5, 10.0],
    ]
    
    for col_num, machine in enumerate(machines, 2):
        for row_num, value in enumerate(machine, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    wb.save('test_files/test_moulding_machines.xlsx')
    print("✓ Generated test_moulding_machines.xlsx (4 moulding machines)")


def generate_assemblies_test():
    """Generate test file for assemblies upload"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Assemblies"
    
    headers = [
        'Assembly Name*', 'Assembly Type', 'Manual Cost',
        'Other Cost', 'Profit %', 'Rejection %', 'Inspection & Handling %'
    ]
    
    add_vertical_headers(ws, headers, "FFC000")
    
    # Add 3 sample assemblies
    assemblies = [
        ['Manual Screw Assembly', 'Manual', 12.50, 3.50, 10.0, 2.5, 3.0],
        ['Ultrasonic Welding', 'Automated', 25.00, 5.00, 15.0, 3.0, 4.0],
        ['Pad Printing', 'Semi-Auto', 8.75, 2.25, 12.0, 2.0, 2.5],
    ]
    
    for col_num, assembly in enumerate(assemblies, 2):
        for row_num, value in enumerate(assembly, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    wb.save('test_files/test_assemblies.xlsx')
    print("✓ Generated test_assemblies.xlsx (3 assemblies)")


def generate_packaging_test():
    """Generate test file for packaging upload"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Packaging"
    
    headers = [
        'Packaging Type*', 'Packaging Length (mm)*', 'Packaging Breadth (mm)*',
        'Packaging Height (mm)*', 'Polybag Length*', 'Polybag Width*',
        'Lifecycle*', 'Cost*', 'Maintenance %', 'Parts per Polybag*'
    ]
    
    add_vertical_headers(ws, headers, "E26B0A")
    
    # Add 3 sample packaging
    packagings = [
        ['pp_box', 600, 400, 250, 16, 20, 150, 8.50, 5.0, 50],
        ['cg_box', 700, 450, 300, 18, 22, 200, 15.00, 4.0, 100],
        ['bin', 800, 500, 350, 20, 24, 500, 45.00, 3.5, 200],
    ]
    
    for col_num, packaging in enumerate(packagings, 2):
        for row_num, value in enumerate(packaging, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    wb.save('test_files/test_packaging.xlsx')
    print("✓ Generated test_packaging.xlsx (3 packaging options)")


def generate_transport_test():
    """Generate test file for transport upload"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transport"
    
    headers = [
        'Transport Length*', 'Transport Breadth*', 'Transport Height*',
        'Trip Cost*', 'Parts per Box*'
    ]
    
    add_vertical_headers(ws, headers, "9933FF")
    
    # Add 2 sample transports
    transports = [
        [3600, 2400, 1800, 6500.00, 120],
        [4200, 2800, 2100, 8500.00, 200],
    ]
    
    for col_num, transport in enumerate(transports, 2):
        for row_num, value in enumerate(transport, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    wb.save('test_files/test_transport.xlsx')
    print("✓ Generated test_transport.xlsx (2 transport options)")


def generate_complete_quote_test():
    """Generate test file for complete quote upload"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Quote Definition (not needed for single quote upload, but good to have)
    ws_def = wb.create_sheet("Quote Definition")
    def_headers = [
        'Quote Name*', 'Client Name*', 'SAP Number', 'Part Number*', 'Part Name*',
        'Amendment Number', 'Description', 'Quantity*', 'Handling Charge', 'Profit %', 'Notes'
    ]
    add_vertical_headers(ws_def, def_headers, "2E75B6")
    
    # Single quote data
    quote_data = ['Test Complete Quote', 'Sample Client Corp', 'SAP-TEST-001', 'PART-ABC-123', 
                  'Widget Assembly Type A', 'Rev-A', 'Complete test quote with all components',
                  5000, 250.00, 18.0, 'Generated for testing']
    
    for row_num, value in enumerate(quote_data, 1):
        ws_def.cell(row=row_num, column=2, value=value)
    ws_def.column_dimensions['B'].width = 25
    
    # Raw Materials Sheet
    ws_rm = wb.create_sheet("Raw Materials")
    rm_headers = [
        'Material Name*', 'Grade', 'RM Code*', 'Unit (kg/gm/ton)*',
        'RM Rate*', 'Frozen Rate', 'Part Weight*', 'Runner Weight*',
        'Process Losses', 'Purging Loss Cost', 'ICC %',
        'Rejection %', 'Overhead %', 'Maintenance %', 'Profit %'
    ]
    add_vertical_headers(ws_rm, rm_headers, "4472C4")
    
    materials = [
        ['Polypropylene', 'PP-H340R', 'RM-PP-001', 'kg', 125.50, None, 0.0234, 0.0045, 2.5, 1.2, 0.5, 2.0, 5.0, 3.0, 10.0],
        ['ABS Resin', 'ABS-750', 'RM-ABS-002', 'kg', 185.75, None, 0.0456, 0.0089, 3.0, 1.5, 0.75, 2.5, 4.5, 2.5, 12.0],
    ]
    
    for col_num, material in enumerate(materials, 2):
        for row_num, value in enumerate(material, 1):
            ws_rm.cell(row=row_num, column=col_num, value=value)
        ws_rm.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Moulding Machines Sheet
    ws_mm = wb.create_sheet("Moulding Machines")
    mm_headers = [
        'Cavity*', 'Machine Tonnage*', 'Cycle Time (s)*', 'Efficiency %*',
        'Shift Rate*', 'Shift Rate for MTC*', 'MTC Count*',
        'Rejection %', 'Overhead %', 'Maintenance %', 'Profit %'
    ]
    add_vertical_headers(ws_mm, mm_headers, "70AD47")
    
    machines = [
        [4, 180, 45.2, 90.0, 6000.00, 5500.00, 3, 2.0, 4.5, 3.0, 12.0],
    ]
    
    for col_num, machine in enumerate(machines, 2):
        for row_num, value in enumerate(machine, 1):
            ws_mm.cell(row=row_num, column=col_num, value=value)
        ws_mm.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Assemblies Sheet
    ws_asm = wb.create_sheet("Assemblies")
    asm_headers = [
        'Assembly Name*', 'Assembly Type', 'Manual Cost',
        'Other Cost', 'Profit %', 'Rejection %', 'Inspection & Handling %'
    ]
    add_vertical_headers(ws_asm, asm_headers, "FFC000")
    
    assemblies = [
        ['Manual Screw Assembly', 'Manual', 12.50, 3.50, 10.0, 2.5, 3.0],
    ]
    
    for col_num, assembly in enumerate(assemblies, 2):
        for row_num, value in enumerate(assembly, 1):
            ws_asm.cell(row=row_num, column=col_num, value=value)
        ws_asm.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Packaging Sheet
    ws_pkg = wb.create_sheet("Packaging")
    pkg_headers = [
        'Packaging Type*', 'Packaging Length (mm)*', 'Packaging Breadth (mm)*',
        'Packaging Height (mm)*', 'Polybag Length*', 'Polybag Width*',
        'Lifecycle*', 'Cost*', 'Maintenance %', 'Parts per Polybag*'
    ]
    add_vertical_headers(ws_pkg, pkg_headers, "E26B0A")
    
    packagings = [
        ['pp_box', 600, 400, 250, 16, 20, 150, 8.50, 5.0, 50],
    ]
    
    for col_num, packaging in enumerate(packagings, 2):
        for row_num, value in enumerate(packaging, 1):
            ws_pkg.cell(row=row_num, column=col_num, value=value)
        ws_pkg.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Transport Sheet
    ws_trans = wb.create_sheet("Transport")
    trans_headers = [
        'Transport Length*', 'Transport Breadth*', 'Transport Height*',
        'Trip Cost*', 'Parts per Box*'
    ]
    add_vertical_headers(ws_trans, trans_headers, "9933FF")
    
    transports = [
        [3600, 2400, 1800, 6500.00, 120],
    ]
    
    for col_num, transport in enumerate(transports, 2):
        for row_num, value in enumerate(transport, 1):
            ws_trans.cell(row=row_num, column=col_num, value=value)
        ws_trans.column_dimensions[get_column_letter(col_num)].width = 15
    
    wb.save('test_files/test_complete_quote.xlsx')
    print("✓ Generated test_complete_quote.xlsx (1 complete quote with all components)")


def generate_multiple_quotes_test():
    """Generate test file for multiple quotes upload"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Instructions Sheet
    ws_inst = wb.create_sheet("Instructions", 0)
    ws_inst.column_dimensions['A'].width = 100
    
    instructions = [
        "TEST FILE - MULTIPLE COMPLETE QUOTES",
        "",
        "This file contains 3 complete quotes with all components:",
        "1. Automotive Dashboard Component",
        "2. Electronic Housing",
        "3. Medical Device Part",
        "",
        "Each quote has:",
        "- Quote definition",
        "- 2-3 raw materials",
        "- 1-2 moulding machines",
        "- 1 assembly",
        "- 1 packaging",
        "- 1 transport",
        "",
        "Upload this to a project to test bulk quote creation.",
    ]
    
    for row_num, instruction in enumerate(instructions, 1):
        cell = ws_inst.cell(row=row_num, column=1)
        cell.value = instruction
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
    
    # Quote Definition Sheet
    ws_def = wb.create_sheet("Quote Definition")
    def_headers = [
        'Quote Name*', 'Client Name*', 'SAP Number', 'Part Number*', 'Part Name*',
        'Amendment Number', 'Description', 'Quantity*', 'Handling Charge', 'Profit %', 'Notes'
    ]
    add_vertical_headers(ws_def, def_headers, "2E75B6")
    
    quotes = [
        ['Auto Dashboard 2024', 'AutoTech Industries', 'SAP-AUTO-2024-001', 'DASH-4567', 
         'Dashboard Center Console', 'Rev-B', 'Automotive dashboard component with premium finish',
         10000, 500.00, 20.0, 'High volume production'],
        ['Electronics Housing Pro', 'TechCorp Ltd', 'SAP-TECH-2024-002', 'EH-PRO-890',
         'Premium Electronics Housing', 'Rev-A', 'IP67 rated electronic component housing',
         25000, 350.00, 18.0, 'Weather resistant coating required'],
        ['Medical Device Shell', 'MedEquip Co', 'SAP-MED-2024-003', 'MD-SHELL-123',
         'Sterile Medical Device Shell', 'Rev-C', 'FDA compliant medical device component',
         5000, 800.00, 25.0, 'Cleanroom manufacturing required'],
    ]
    
    for col_num, quote_data in enumerate(quotes, 2):
        for row_num, value in enumerate(quote_data, 1):
            ws_def.cell(row=row_num, column=col_num, value=value)
        ws_def.column_dimensions[get_column_letter(col_num)].width = 20
    
    # Raw Materials Sheet
    ws_rm = wb.create_sheet("Raw Materials")
    rm_headers = [
        'Quote Name*', 'Material Name*', 'Grade', 'RM Code*', 'Unit (kg/gm/ton)*',
        'RM Rate*', 'Frozen Rate', 'Part Weight*', 'Runner Weight*',
        'Process Losses', 'Purging Loss Cost', 'ICC %',
        'Rejection %', 'Overhead %', 'Maintenance %', 'Profit %'
    ]
    add_vertical_headers(ws_rm, rm_headers, "4472C4")
    
    # Raw materials for each quote
    raw_materials = [
        # Quote 1 materials
        ['Auto Dashboard 2024', 'PP Copolymer', 'PP-R272', 'RM-PP-272', 'kg', 135.50, None, 0.245, 0.038, 3.5, 1.8, 0.8, 2.2, 5.5, 3.2, 11.0],
        ['Auto Dashboard 2024', 'ABS High Impact', 'ABS-HI-850', 'RM-ABS-850', 'kg', 195.00, None, 0.156, 0.024, 2.8, 1.5, 0.6, 2.0, 5.0, 3.0, 10.0],
        ['Auto Dashboard 2024', 'TPE Soft Touch', 'TPE-ST-60', 'RM-TPE-060', 'kg', 245.00, 240.00, 0.089, 0.012, 1.5, 1.0, 0.4, 1.8, 4.5, 2.8, 9.0],
        # Quote 2 materials
        ['Electronics Housing Pro', 'PC+ABS Blend', 'PCABS-750', 'RM-PCABS-750', 'kg', 265.00, None, 0.178, 0.028, 3.2, 1.6, 0.7, 2.5, 6.0, 3.5, 12.0],
        ['Electronics Housing Pro', 'Glass Filled PA6', 'PA6-GF30', 'RM-PA6-030', 'kg', 285.00, 280.00, 0.234, 0.036, 4.0, 2.0, 1.0, 3.0, 6.5, 4.0, 15.0],
        # Quote 3 materials
        ['Medical Device Shell', 'Medical Grade PP', 'PP-MED-H310', 'RM-PP-MED-310', 'kg', 385.00, None, 0.125, 0.019, 2.0, 1.2, 0.5, 1.5, 4.0, 2.5, 8.0],
        ['Medical Device Shell', 'Medical Grade PC', 'PC-MED-1800', 'RM-PC-MED-1800', 'kg', 495.00, 490.00, 0.167, 0.025, 2.5, 1.5, 0.6, 1.8, 4.5, 2.8, 10.0],
    ]
    
    for col_num, material in enumerate(raw_materials, 2):
        for row_num, value in enumerate(material, 1):
            ws_rm.cell(row=row_num, column=col_num, value=value)
        ws_rm.column_dimensions[get_column_letter(col_num)].width = 18
    
    # Moulding Machines Sheet
    ws_mm = wb.create_sheet("Moulding Machines")
    mm_headers = [
        'Quote Name*', 'Cavity*', 'Machine Tonnage*', 'Cycle Time (s)*', 'Efficiency %*',
        'Shift Rate*', 'Shift Rate for MTC*', 'MTC Count*',
        'Rejection %', 'Overhead %', 'Maintenance %', 'Profit %'
    ]
    add_vertical_headers(ws_mm, mm_headers, "70AD47")
    
    machines = [
        # Quote 1 machines
        ['Auto Dashboard 2024', 8, 320, 58.5, 88.0, 9500.00, 9000.00, 4, 2.8, 6.5, 4.2, 16.0],
        ['Auto Dashboard 2024', 4, 220, 52.3, 90.0, 7200.00, 6800.00, 3, 2.5, 6.0, 3.8, 14.0],
        # Quote 2 machines
        ['Electronics Housing Pro', 4, 180, 45.7, 91.0, 6500.00, 6100.00, 3, 2.2, 5.5, 3.5, 13.0],
        # Quote 3 machines
        ['Medical Device Shell', 2, 150, 42.8, 92.5, 5800.00, 5400.00, 2, 2.0, 5.0, 3.2, 12.0],
    ]
    
    for col_num, machine in enumerate(machines, 2):
        for row_num, value in enumerate(machine, 1):
            ws_mm.cell(row=row_num, column=col_num, value=value)
        ws_mm.column_dimensions[get_column_letter(col_num)].width = 18
    
    # Assemblies Sheet
    ws_asm = wb.create_sheet("Assemblies")
    asm_headers = [
        'Quote Name*', 'Assembly Name*', 'Assembly Type', 'Manual Cost',
        'Other Cost', 'Profit %', 'Rejection %', 'Inspection & Handling %'
    ]
    add_vertical_headers(ws_asm, asm_headers, "FFC000")
    
    assemblies = [
        ['Auto Dashboard 2024', 'Insert Molding + Ultrasonic Welding', 'Automated', 28.50, 6.50, 12.0, 3.0, 4.5],
        ['Electronics Housing Pro', 'Snap Fit Assembly', 'Manual', 15.75, 4.25, 10.0, 2.5, 3.5],
        ['Medical Device Shell', 'Cleanroom Assembly + Sterilization', 'Manual', 45.00, 12.00, 15.0, 1.5, 5.0],
    ]
    
    for col_num, assembly in enumerate(assemblies, 2):
        for row_num, value in enumerate(assembly, 1):
            ws_asm.cell(row=row_num, column=col_num, value=value)
        ws_asm.column_dimensions[get_column_letter(col_num)].width = 18
    
    # Packaging Sheet
    ws_pkg = wb.create_sheet("Packaging")
    pkg_headers = [
        'Quote Name*', 'Packaging Type*', 'Packaging Length (mm)*', 'Packaging Breadth (mm)*',
        'Packaging Height (mm)*', 'Polybag Length*', 'Polybag Width*',
        'Lifecycle*', 'Cost*', 'Maintenance %', 'Parts per Polybag*'
    ]
    add_vertical_headers(ws_pkg, pkg_headers, "E26B0A")
    
    packagings = [
        ['Auto Dashboard 2024', 'pp_box', 700, 450, 300, 18, 22, 200, 12.50, 4.5, 80],
        ['Electronics Housing Pro', 'cg_box', 600, 400, 250, 16, 20, 250, 18.00, 3.8, 100],
        ['Medical Device Shell', 'bin', 500, 350, 200, 14, 18, 500, 65.00, 2.5, 25],
    ]
    
    for col_num, packaging in enumerate(packagings, 2):
        for row_num, value in enumerate(packaging, 1):
            ws_pkg.cell(row=row_num, column=col_num, value=value)
        ws_pkg.column_dimensions[get_column_letter(col_num)].width = 18
    
    # Transport Sheet
    ws_trans = wb.create_sheet("Transport")
    trans_headers = [
        'Quote Name*', 'Transport Length*', 'Transport Breadth*', 'Transport Height*',
        'Trip Cost*', 'Parts per Box*'
    ]
    add_vertical_headers(ws_trans, trans_headers, "9933FF")
    
    transports = [
        ['Auto Dashboard 2024', 4200, 2800, 2100, 9500.00, 240],
        ['Electronics Housing Pro', 3600, 2400, 1800, 7200.00, 300],
        ['Medical Device Shell', 3000, 2000, 1500, 5500.00, 100],
    ]
    
    for col_num, transport in enumerate(transports, 2):
        for row_num, value in enumerate(transport, 1):
            ws_trans.cell(row=row_num, column=col_num, value=value)
        ws_trans.column_dimensions[get_column_letter(col_num)].width = 18
    
    wb.save('test_files/test_multiple_quotes.xlsx')
    print("✓ Generated test_multiple_quotes.xlsx (3 complete quotes with all components)")


def main():
    """Generate all test files"""
    print("\n" + "="*60)
    print("Generating Test Excel Files for Quote Management System")
    print("="*60 + "\n")
    
    create_test_files_directory()
    print()
    
    print("Generating component test files...")
    generate_raw_materials_test()
    generate_moulding_machines_test()
    generate_assemblies_test()
    generate_packaging_test()
    generate_transport_test()
    print()
    
    print("Generating complete quote test files...")
    generate_complete_quote_test()
    generate_multiple_quotes_test()
    print()
    
    print("="*60)
    print("✓ All test files generated successfully!")
    print("="*60)
    print("\nGenerated files in 'test_files/' directory:")
    print("  1. test_raw_materials.xlsx - 5 raw materials")
    print("  2. test_moulding_machines.xlsx - 4 moulding machines")
    print("  3. test_assemblies.xlsx - 3 assemblies")
    print("  4. test_packaging.xlsx - 3 packaging options")
    print("  5. test_transport.xlsx - 2 transport options")
    print("  6. test_complete_quote.xlsx - 1 complete quote")
    print("  7. test_multiple_quotes.xlsx - 3 complete quotes")
    print("\nYou can now upload these files to test the system!")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()
